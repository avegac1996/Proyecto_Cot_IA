"""Generación de cotizaciones en PDF y Excel."""

import io
import os
from datetime import datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image as RLImage,
)

from app.models.cotizacion import Cotizacion

# Brand colors
BRAND_PRIMARY = colors.HexColor("#06b6d4")
BRAND_PRIMARY_DARK = colors.HexColor("#0891b2")
BRAND_PRIMARY_LIGHT = colors.HexColor("#cffafe")
BRAND_BG_LIGHT = colors.HexColor("#f0f9ff")
BRAND_TEXT = colors.HexColor("#0f172a")
BRAND_TEXT_MUTED = colors.HexColor("#64748b")
BRAND_BORDER = colors.HexColor("#e2e8f0")
BRAND_DANGER = colors.HexColor("#ef4444")
BRAND_SUCCESS = colors.HexColor("#10b981")

LOGO_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "static", "logo.png")


def _money(value: Decimal | float | int | None) -> str:
    if value is None:
        return "—"
    return f"${Decimal(str(value)):.2f}"


def generate_pdf(cotizacion: Cotizacion, iva_pct: float = 15.0) -> bytes:
    """Genera un PDF de la cotización y retorna los bytes."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )

    styles = getSampleStyleSheet()

    # --- Styles ---
    brand_title = ParagraphStyle(
        "BrandTitle",
        parent=styles["Normal"],
        fontSize=22,
        fontName="Helvetica-Bold",
        textColor=BRAND_TEXT,
        spaceAfter=2,
    )
    brand_subtitle = ParagraphStyle(
        "BrandSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=BRAND_TEXT_MUTED,
        spaceAfter=0,
    )
    info_label = ParagraphStyle(
        "InfoLabel",
        parent=styles["Normal"],
        fontSize=8,
        fontName="Helvetica-Bold",
        textColor=BRAND_TEXT_MUTED,
        spaceAfter=1,
    )
    info_value = ParagraphStyle(
        "InfoValue",
        parent=styles["Normal"],
        fontSize=10,
        textColor=BRAND_TEXT,
        spaceAfter=0,
    )
    section_style = ParagraphStyle(
        "Section",
        parent=styles["Normal"],
        fontSize=11,
        fontName="Helvetica-Bold",
        textColor=BRAND_PRIMARY_DARK,
        spaceBefore=16,
        spaceAfter=8,
    )
    cell_style = ParagraphStyle(
        "Cell",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        textColor=BRAND_TEXT,
    )
    cell_center = ParagraphStyle("CellCenter", parent=cell_style, alignment=1)
    cell_right = ParagraphStyle("CellRight", parent=cell_style, alignment=2)
    footer_style = ParagraphStyle(
        "Footer",
        parent=styles["Normal"],
        fontSize=8,
        textColor=BRAND_TEXT_MUTED,
        alignment=1,
    )
    total_label_style = ParagraphStyle(
        "TotalLabel",
        parent=cell_right,
        fontName="Helvetica-Bold",
        fontSize=11,
        textColor=BRAND_TEXT,
    )
    total_value_style = ParagraphStyle(
        "TotalValue",
        parent=cell_right,
        fontName="Helvetica-Bold",
        fontSize=13,
        textColor=BRAND_PRIMARY_DARK,
    )

    elements = []

    # --- Header with logo ---
    logo_flow = None
    if os.path.exists(LOGO_PATH):
        try:
            logo_flow = RLImage(LOGO_PATH, width=1.4 * inch, height=1.4 * inch, kind="proportional")
        except Exception:
            logo_flow = None

    header_left = []
    header_left.append(Spacer(1, 4))
    header_left.append(Paragraph("Sistema Inteligente de Cotización", brand_subtitle))
    header_left.append(Paragraph("Componentes Electrónicos", brand_subtitle))

    if logo_flow:
        header_table = Table(
            [[logo_flow, header_left]],
            colWidths=[1.8 * inch, 5.0 * inch],
        )
        header_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (0, 0), 16),
            ("RIGHTPADDING", (1, 0), (1, 0), 0),
        ]))
        elements.append(header_table)
    else:
        elements.extend(header_left)

    # Color bar separator
    bar = Table([[""]], colWidths=[7.0 * inch], rowHeights=[4])
    bar.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), BRAND_PRIMARY)]))
    elements.append(Spacer(1, 8))
    elements.append(bar)
    elements.append(Spacer(1, 12))

    # --- Client data ---
    if cotizacion.cliente_nombre or cotizacion.cliente_correo or cotizacion.cliente_celular:
        cliente_lines = []
        if cotizacion.cliente_nombre:
            cliente_lines.append(Paragraph(
                f"<b>Cliente:</b> {cotizacion.cliente_nombre}",
                ParagraphStyle("ClienteData", parent=styles["Normal"], fontSize=10, textColor=BRAND_TEXT, spaceAfter=2),
            ))
        if cotizacion.cliente_correo:
            cliente_lines.append(Paragraph(
                f"<b>Correo:</b> {cotizacion.cliente_correo}",
                ParagraphStyle("ClienteData2", parent=styles["Normal"], fontSize=10, textColor=BRAND_TEXT, spaceAfter=2),
            ))
        if cotizacion.cliente_celular:
            cliente_lines.append(Paragraph(
                f"<b>Celular:</b> {cotizacion.cliente_celular}",
                ParagraphStyle("ClienteData3", parent=styles["Normal"], fontSize=10, textColor=BRAND_TEXT, spaceAfter=2),
            ))
        cliente_box = Table([[cliente_lines]], colWidths=[7.0 * inch])
        cliente_box.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), BRAND_BG_LIGHT),
            ("BOX", (0, 0), (-1, -1), 0.5, BRAND_BORDER),
            ("LEFTPADDING", (0, 0), (-1, -1), 12),
            ("RIGHTPADDING", (0, 0), (-1, -1), 12),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(cliente_box)
        elements.append(Spacer(1, 16))

    # --- Items table ---

    def _p(text: str, style: ParagraphStyle = cell_style) -> Paragraph:
        escaped = (text or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        return Paragraph(escaped, style)

    hdr_style = ParagraphStyle("Hdr", parent=cell_style, fontName="Helvetica-Bold", textColor=colors.white, fontSize=9)
    hdr_center = ParagraphStyle("HdrC", parent=cell_center, fontName="Helvetica-Bold", textColor=colors.white, fontSize=9)
    hdr_right = ParagraphStyle("HdrR", parent=cell_right, fontName="Helvetica-Bold", textColor=colors.white, fontSize=9)

    header = [
        Paragraph("Producto", hdr_style),
        Paragraph("Cant.", hdr_center),
        Paragraph("P. Unit.", hdr_right),
        Paragraph("Subtotal", hdr_right),
    ]
    rows = [header]
    for item in cotizacion.items:
        rows.append([
            _p(item.producto_nombre),
            _p(str(item.cantidad), cell_center),
            _p(_money(item.precio_unitario) if item.disponible else "—", cell_right),
            _p(_money(item.subtotal) if item.disponible else "—", cell_right),
        ])

    # Subtotal, IVA, Total rows
    subtotal = cotizacion.total
    iva_amount = (subtotal * Decimal(str(iva_pct)) / Decimal("100")).quantize(Decimal("0.01"))
    total_con_iva = (subtotal + iva_amount).quantize(Decimal("0.01"))

    rows.append([
        "", "",
        Paragraph("Subtotal:", total_label_style),
        Paragraph(_money(subtotal), total_value_style),
    ])
    rows.append([
        "", "",
        Paragraph(f"IVA ({iva_pct:.0f}%):", total_label_style),
        Paragraph(_money(iva_amount), total_value_style),
    ])
    rows.append([
        "", "",
        Paragraph("TOTAL:", total_label_style),
        Paragraph(_money(total_con_iva), total_value_style),
    ])

    col_widths = [3.5 * inch, 0.7 * inch, 1.2 * inch, 1.2 * inch]
    table = Table(rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Header row
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_PRIMARY_DARK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        # Body rows
        ("ROWBACKGROUNDS", (0, 1), (-1, -4), [colors.white, BRAND_BG_LIGHT]),
        ("GRID", (0, 0), (-1, -4), 0.4, BRAND_BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 1), (-1, -4), 6),
        ("BOTTOMPADDING", (0, 1), (-1, -4), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        # Summary rows (subtotal, iva, total)
        ("LINEABOVE", (0, -3), (-1, -3), 1, BRAND_PRIMARY),
        ("TOPPADDING", (0, -3), (-1, -3), 8),
        ("BOTTOMPADDING", (0, -3), (-1, -3), 4),
        ("TOPPADDING", (0, -2), (-1, -2), 4),
        ("BOTTOMPADDING", (0, -2), (-1, -2), 4),
        # Total row
        ("BACKGROUND", (0, -1), (-1, -1), BRAND_PRIMARY_LIGHT),
        ("LINEABOVE", (0, -1), (-1, -1), 2, BRAND_PRIMARY),
        ("TOPPADDING", (0, -1), (-1, -1), 10),
        ("BOTTOMPADDING", (0, -1), (-1, -1), 10),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 30))

    # --- Footer ---
    bar2 = Table([[""]], colWidths=[7.0 * inch], rowHeights=[2])
    bar2.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), BRAND_PRIMARY)]))
    elements.append(bar2)
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(
        "AV Electronics • Sucursal Almagro: Andrade Marín e7-76 y Av Diego de Almagro • Teléfono: 0999 200 997 • E-mail: ventas@avelectronics.cc • Quito - Ecuador",
        footer_style,
    ))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def generate_excel(cotizacion: Cotizacion) -> bytes:
    """Genera un archivo Excel de la cotización y retorna los bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Cotización #{cotizacion.id}"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=16, color="0891B2")
    subtitle_font = Font(name="Calibri", size=10, color="64748B")
    total_font = Font(name="Calibri", bold=True, size=12, color="0891B2")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "AV Electronics — Cotización de Componentes Electrónicos"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    fecha_str = cotizacion.fecha_creacion.strftime("%d/%m/%Y %H:%M") if cotizacion.fecha_creacion else ""
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Cotización #{cotizacion.id} · {fecha_str} · Estado: {cotizacion.estado.upper()}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers (row 4)
    headers = ["Producto", "Cantidad", "Proveedor", "P. Unitario", "Subtotal", "Estado"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Items
    row = 5
    for item in cotizacion.items:
        ws.cell(row=row, column=1, value=item.producto_nombre).border = thin_border
        ws.cell(row=row, column=2, value=item.cantidad).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=item.proveedor or "—").border = thin_border
        if item.disponible:
            ws.cell(row=row, column=4, value=float(item.precio_unitario)).border = thin_border
            ws.cell(row=row, column=4).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=5, value=float(item.subtotal)).border = thin_border
            ws.cell(row=row, column=5).number_format = '"$"#,##0.00'
        else:
            ws.cell(row=row, column=4, value="—").border = thin_border
            ws.cell(row=row, column=5, value="—").border = thin_border
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
        estado_cell = ws.cell(row=row, column=6, value="Disponible" if item.disponible else "Sin datos")
        estado_cell.border = thin_border
        estado_cell.alignment = Alignment(horizontal="center")
        row += 1

    # Total row
    ws.cell(row=row, column=4, value="TOTAL:").font = total_font
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=row, column=5, value=float(cotizacion.total))
    total_cell.font = total_font
    total_cell.number_format = '"$"#,##0.00'
    total_cell.alignment = Alignment(horizontal="right")

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes
